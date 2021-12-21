from PIL import Image,ImageDraw,ImageFont
import openpyxl
from openpyxl import load_workbook
import smtplib as s




def xls_to_pylist(filepath,template_path,font_path,font_size,ext):
    
     

    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    
    no_of_row_in_sheet=8

    for i in range(2,no_of_row_in_sheet):

        get_name=(sheet.cell(row=i,column=1))
        certi_name=get_name.value
        get_email=(sheet.cell(row=i,column=2))
        certi_email=get_email.value
        print(certi_name)
        print(certi_email)
        
 

    
        img=Image.open(template_path,"r")
        #img.show()
        
        draw=ImageDraw.Draw(img)
        font=ImageFont.truetype(font_path,size=font_size)
        color="rgb(0,0,0)"
        x,y=(100,80)
        draw.text((x,y),certi_name,fill=color,font=font)


        # Save images

        saving_path=f"/home/vaishu/Desktop/python pro/certificate generator/certificates/{certi_name}"
        img.save(saving_path,ext)
        
    
        
       
xls_to_pylist("/home/vaishu/Desktop/python pro/certificate generator/appoinment letters.xlsx","/home/vaishu/Desktop/python pro/certificate generator/download.jpeg","arial.ttf",10,"JPEG")
